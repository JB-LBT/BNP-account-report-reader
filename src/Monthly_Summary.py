###########################################################################################

################## This package has been written by JB LBT (c) 2024 #######################
################## Under the GNU GPL v3.0 Licence                   #######################

###########################################################################################


# Importing the necessary libraries
import os
import nest_asyncio
import os
import pandas as pd
from functools import wraps
from datetime import datetime
from llama_parse import LlamaParse
from pickle import dump, load
#from dotenv import load_dotenv
from src.utils import *

with open("llamaparse_key.txt", "r") as f:
    os.environ["LLAMA_CLOUD_API_KEY"] = f.read().strip()

nest_asyncio.apply()

# read the .env file
#load_dotenv()

months_mapping = {"janvier":1, "février":2, "mars":3, "avril":4, "mai":5, "juin":6, "juillet":7, "août":8, "septembre":9, "octobre":10, "novembre":11, "décembre":12}
months_mapping_inv = {v: k for k, v in months_mapping.items()}
#ACCOUNT_ID = os.getenv("ACCOUNT_ID")
ACCOUNT_ID = "JB_courant"
CATEGORY_LIST = ["Transports", "Vie quotidienne", "Logement", "Loisirs", "Santé", "Impôts", "Banque", "Salaire", "Epargne", "Autre"]
LOG_FOLDER = "logs"
if not os.path.exists(LOG_FOLDER):
    os.makedirs(LOG_FOLDER)




############################################################################################

##################################### CORE FUNCTIONS #######################################

############################################################################################




class Statement_Parser:
    def __init__(self, 
                 document,
                 mode = "text",
                 logger = None,
                 **kwargs):
        self.document = document
        self.mode = mode
        self.parse_args(kwargs)
        if not logger:
            now = datetime.now()
            # format the date as YYYY-MM-DD_HH-MM-SS
            now = now.strftime("%Y-%m-%d_%H-%M-%S")
            logger_name = f"{LOG_FOLDER}/parser_{self.pdf.split('/')[-1].split('.')[0]}_{now}.log"
            self.logger = Logger(logger_name,self.formatting, self.do_log)
        else:
            self.logger = logger

        self.logger.log("Parser initialized")
        self.logger.log(f"Document to parse: {self.document}")
        self.logger.log(f"Mode: {self.mode}")


    def parse_args(self, kwargs):
        """Parse the keyword arguments and set the attributes of the class.

        Parameters:
        -----------
        kwargs : dict
            The keyword arguments to parse.

        Returns:
        --------
        None
        """

        for key, value in kwargs.items():
            setattr(self, key, value)

        if "verbose" not in kwargs and not hasattr(self, "verbose"):
            self.verbose = True
        if "do_log" not in kwargs and not hasattr(self, "do_log"):
            self.do_log = True
        if "formatting" not in kwargs and not hasattr(self, "formatting"):
            self.formatting = True
        if "rules_file" not in kwargs and not hasattr(self, "rules_file"):
            self.rules_file = None
        if "ask_rules" not in kwargs and not hasattr(self, "ask_rules"):
            self.ask_rules = False
        if "handle_errors" not in kwargs and not hasattr(self, "handle_errors"):
            self.handle_errors = True

        return None
    
    def format_dataframe(self):
        if not hasattr(self, "data"):
            if self.handle_errors:
                self.logger.warning("No data to format")
                return None
            else:
                self.logger.error("No data to format", title="ERROR when formatting parser data")
                raise ValueError("No data to format")
            
        try:            
            # Fill the missing values and blanks " " with 0
            self.data["Debit (€)"] = self.data["Debit (€)"].str.replace(" ", "")
            self.data["Credit (€)"] = self.data["Credit (€)"].str.replace(" ", "")
            
            # Fill the empty values with 0
            self.data["Debit (€)"] = self.data["Debit (€)"].replace("", "0")
            self.data["Credit (€)"] = self.data["Credit (€)"].replace("", "0")
            
            # Convert the amounts to float
            self.data["Debit (€)"] = self.data["Debit (€)"].str.replace(",", ".").astype(float)
            self.data["Credit (€)"] = self.data["Credit (€)"].str.replace(",", ".").astype(float)

        except Exception as e:
            self.logger.error(f"Error while formatting the operations amounts: {e}", title = "Formatting error")
            if not self.handle_errors:
                raise e
            
        try:
            # check if the dates are between december and january
            december = False
            january = False
            for i, date in enumerate(self.data["Date"]):
                if date.split(".")[1] == "12":
                    december = True
                if date.split(".")[1] == "01":
                    january = True

            def add_date(date):
                if date.split(".")[1] == "01" and december and january:
                    return date + f".{self.start_year+1}"
                else:
                    return date + f".{self.start_year}"
                
            self.data["Date"] = self.data["Date"].apply(add_date)
            self.data["Operation Date"] = self.data["Operation Date"].apply(add_date)

            # Convert the dates to datetime
            self.data["Date"] = pd.to_datetime(self.data["Date"], format="%d.%m.%Y")
            self.data["Operation Date"] = pd.to_datetime(self.data["Operation Date"], format="%d.%m.%Y")

        except Exception as e:
            self.logger.error(f"Error while formatting the operations dates: {e}", title = "Formatting error")
            if not self.handle_errors:
                raise e
            
        return None
    
    def load_document(self):
        """Load the PDF file and parse it to extract the operations."""
        from llama_parse import LlamaParse
        parser = LlamaParse( # can also be set in your env as LLAMA_CLOUD_API_KEY
            result_type=self.mode,  # "markdown" and "text" are available
            num_workers=4,  # if multiple files passed, split in `num_workers` API calls
            verbose=False,
            language="en",  # Optionally you can define a language, default=en
        )

        self.parsed_document = None

        pickle_file = f"{self.document.split('.')[0]}_{self.mode}.pkl"
        if os.path.exists(pickle_file):
            try:
                with open(pickle_file, "rb") as f:
                    self.parsed_document = load(f)
                    self.logger.log(f"Document {self.document} successfully loaded from pickle file {pickle_file}")
                    return None
            except Exception as e:
                self.logger.error(f"Error while loading the pickle file: {e}", title = "Loading error")
                self.logger.warning("Deleting the pickle file and parsing the document from scratch")
                os.remove(pickle_file)
                self.logger.warning("Parsing the document from scratch")

        # sync
        try:
            documents = parser.load_data(self.document)
        except Exception as e:
            self.logger.error(f"Error while loading the document: {e}", title = "Loading error")
            if not self.handle_errors:
                raise e
            
        self.logger.log(f"Document successfully loaded in mode {self.mode}")
        self.parsed_document = documents
        dump(documents, open(pickle_file, "wb"))
        self.logger.log(f"Document successfully saved to pickle file {pickle_file}")
        return None
    
    def _check_operations(self, page):
        if "Crédit" and "Débit" in page:
            return True

        return False
    
    def _strip_first_page(self, page):

        try:
            for i, line in enumerate(page.split("\n")):
                if "RELEVE DE COMPTE" in line:
                    break
            line = line.split("du")[1].split("au")[0].strip()[-4:]
            self.start_year = int(line)
        except Exception as e:
            self.logger.error(f"Error while getting the start year: {e}", title = "Parsing error")
            self.start_year = 2000
            if not self.handle_errors:
                raise e
            
        try:
            for i, line in enumerate(page.split("\n")):
                if "SOLDE CREDITEUR AU" in line:
                    break
            line = line.split("AU")[1].strip()[:15].strip()
            # Convert from dd.mm.yyyy to datetime
            self.start_date = pd.to_datetime(line, format="%d.%m.%Y")
        except Exception as e:
            self.logger.error(f"Error while getting the start date: {e}", title = "Parsing error")

            if not self.handle_errors:
                raise e 


        if "TOTAL DES OPERATIONS" in page:
            try:
                for i in range(len(page.split("\n"))-1, 0, -1):
                    line = page.split("\n")[i]
                    if "SOLDE CREDITEUR AU" in line:
                        break
                line = line.split("AU")[1].strip()[:15].strip()
                # Convert from dd.mm.yyyy to datetime
                self.end_date = pd.to_datetime(line, format="%d.%m.%Y")
            except Exception as e:
                self.logger.error(f"Error while getting the end date: {e}", title = "Parsing error")

                if not self.handle_errors:
                    raise e 
            page = page.split("TOTAL DES OPERATIONS")[0]
        else:
            page = page.split("BNP PARIBAS SA")[0]

        page = page.split("Monnaie du compte : Euro")[1]
        page_lines = page.split("\n")
        page_lines = [line for line in page_lines if line]
        return page_lines
    
    def _strip_reg_page(self, page):
        if "TOTAL DES OPERATIONS" in page:
            try:
                for i in range(len(page.split("\n"))-1, 0, -1):
                    line = page.split("\n")[i]
                    if "SOLDE CREDITEUR AU" in line:
                        break
                line = line.split("AU")[1].strip()[:15].strip()
                # Convert from dd.mm.yyyy to datetime
                self.end_date = pd.to_datetime(line, format="%d.%m.%Y")
            except Exception as e:
                self.logger.error(f"Error while getting the end date: {e}", title = "Parsing error")

                if not self.handle_errors:
                    raise e 
            page = page.split("TOTAL DES OPERATIONS")[0]
        else:
            page = page.split("\n\n")[-2]
        page_lines = page.split("\n")
        for i, l in enumerate(page_lines):
            if "RIB" in l:
                page_lines = page_lines[i+1:]
                break

        page_lines = [line for line in page_lines if line]
        return page_lines
    
    def _parse_page_txt(self, page_lines):
        data = []
        keys = ["Date", "Description", "Operation Date", "Debit (€)", "Credit (€)"]
        try:
            headers = page_lines[0]
            # find the index of "Nature des opérations"
            desc_index = headers.index("Nature des opérations")
            # find the index of "Valeur"
            valeur_index = headers.index("Valeur")
            # find the index of "Débit"
            debit_index = headers.index("Débit")
            # find the index of "Crédit"
            credit_index = headers.index("Crédit")

        except Exception as e:
            self.logger.error(f"Error while parsing the headers of the operations: {e}", title = "Parsing error")
            if not self.handle_errors:
                raise e

        page_lines = [p for p in page_lines if "SOLDE CREDITEUR" not in p]

        for line in page_lines[1:]:

            if len(line) <= valeur_index:
                data[-1]["Description"] += line.strip()

            else:
                date = line[:desc_index-2].strip()
                description = line[desc_index-2:valeur_index-2].strip()
                valeur = line[valeur_index-2:valeur_index+10].strip()
                if len(line) < credit_index-4:
                    debit = line[valeur_index+10:].strip()
                    debit = debit.replace(" ", "")
                    credit = "0.0"
                else:
                    debit = "0.0"
                    credit = line[valeur_index+10:].strip()
                    credit = credit.replace(" ", "")

                values = [date, description, valeur, debit, credit]
                data.append(dict(zip(keys, values)))

        # convert to dataframe
        df = pd.DataFrame(data)
        return df
    
    def _parse_doc_txt(self):
        df = pd.DataFrame()
        self.logger.log("Parsing document in text mode")
        for i in range(len(self.parsed_document)):
            page = self.parsed_document[i].text
            if self._check_operations(page):
                if i == 0:
                    page_lines = self._strip_first_page(page)
                    self.logger.log(f"Stripping first page {i}")
                else:
                    page_lines = self._strip_reg_page(page)
                    self.logger.log(f"Stripping regular page {i}")

                page_df = self._parse_page_txt(page_lines)
                self.logger.log(f"Page {i} successfully parsed, {len(page_df)} operations found")
                df = pd.concat([df, page_df], axis=0)
            else:
                self.logger.log(f"No operations found in page {i}")

        self.data = df
        self.logger.log(f"Document successfully parsed, {len(df)} operations found")
        return None
    
    def _parse_operations_md(self, doc, page):
        """Parse the operations from a parsed PDF document."""
        d = doc.text
        try:
            info = d.split("|---|---|---|---|---|")[-1]
            info = info.split("\n\n")[0]
            infos = info.split("\n")
        except Exception as e:
            self.logger.error(f"Error parsing the PDF {self.document} on page {page}: {e}", title = "PDF parsing error")
            if not self.handle_errors:
                raise e

        data = []
        keys = ["Date", "Description", "Operation Date", "Debit (€)", "Credit (€)"]
        for i in infos:
            d = i.split("|")[1:6]
            if len(d)==5 and d[0]!="":
                data.append(dict(zip(keys, d)))

        data = pd.DataFrame(data)
        return data
    
    def _parse_doc_md(self):

        try:
            chunks = self.parsed_document[0].text.split("\n")

            # get the first chunk that contains "RELEVE DE COMPTE "
            for date in chunks:
                if "RELEVE DE COMPTE" in date:
                    break
            
            date = date.split("du")[1].split("au")[0].strip()[-4:]
            self.start_year = int(date)
        except Exception as e:
            self.logger.error(f"Error when getting the start year: {e}", title = "Parsing error")
            self.start_year = 2000
            if not self.handle_errors:
                raise e
            

        df = pd.DataFrame(columns = ["Date", "Description", "Operation Date", "Debit (€)", "Credit (€)"])
        self.logger.log("Parsing document in markdown mode")
        if self.parsed_document is not None:
            for page, doc in enumerate(self.parsed_document):
                try:
                    for line in doc.text.split("\n"):
                        if "SOLDE CREDITEUR" in line:
                            break

                    line = line.split("AU")[1].strip()[:15].strip()
                    # Convert from dd.mm.yyyy to datetime
                    if page == 0:
                        self.start_date = pd.to_datetime(line, format="%d.%m.%Y")
                        # Set end date to start date + 1 month
                        self.end_date = self.start_date + pd.DateOffset(months=1)
                    else:
                        self.end_date = pd.to_datetime(line, format="%d.%m.%Y")

                    data = self._parse_operations_md(doc, page)
                    self.logger.log(f"Operations from page {page} successfully parsed")
                except Exception as e:
                    self.logger.error(f"Error parsing the PDF {self.document} on page {page}: {e}", title = "PDF parsing error")
                    if not self.handle_errors:
                        raise e
                    else:
                        continue

                # Add the new data to the existing DataFrame
                df = pd.concat([df, data], ignore_index=True)

                self.logger.log(f"Operations from page {page} successfully added")


        # delete the last 2 lines of the DataFrame as they are not operations
        self.data = df[:-2]
        if len(self.data)>0 and "SOLDE CREDITEUR" in self.data.iloc[0]["Description"]:
            self.data = self.data[1:]

    def parse_document(self):
        if self.mode == "text":
            self._parse_doc_txt()
        else:
            self._parse_doc_md()

        self.format_dataframe()
        return self.data



class Monthly_Summary:
    """
        Monthly_Summary
        ===============
        
        Class to create a monthly summary of bank operations from a PDF file.
        This class contains all the necessary functions to load a PDF file, parse the operations, add categories to the operations and save the data to a CSV or Excel file.
        
        Methods:
        --------
        Here are the main methods of the class:
        
        - add_operations: load the PDF file, parse the operations and add them to the DataFrame
        - modify_operation: modify an operation in the DataFrame
        - plot_categories: plot a pie chart of the categories
        - get_stats: get the statistics of the monthly summary
        - summary: print a summary of the monthly report
        - add_monthly_budget: add a budget to the monthly summary
        - compute_remaining_budget: compute the remaining budget
        - to_csv: save the monthly summary to a CSV file
        - to_excel: save the monthly summary to an Excel file


        Attributes:
        -----------
        - month: int
            The month of the operations.

        - year: int
            The year of the operations.

        - operations: DataFrame
            A DataFrame containing the operations.

        - pdf: str
            The path to the PDF file containing the bank operations.

        - logger: Logger
            A Logger object to log the operations of the class.

        - verbose: int
            The verbosity level of the logger.

        - do_log: bool
            If True, logs will be saved to a file.

        - formatting: bool
            If True, logs will be formatted.

        - rules_file: str
            The path to the rules file.

        - ask_rules: bool
            If True, the user will be asked to provide rules for the categories.

        - handle_errors: bool
            If True, errors will be handled and logged.

        - start_date: datetime
            The start date of the operations.

        - end_date: datetime
            The end date of the operations.

        - budget: float
            The budget of the month.

        - remaining_budget: float
            The remaining budget of the month.

        - CATEGORY_LIST: list
            The list of categories to use for the operations.
        """
    def __init__(self, pdf, month = None, year = None, operations = None, **kwargs):
        """Initialize the Monthly_Summary object.
        
        Parameters:
        -----------
        pdf : str
            The path to the PDF file containing the bank operations.
            
        month : int, optional
            The month of the operations. If not provided, the month will be extracted from the PDF file.
            
        year : int, optional
            The year of the operations. If not provided, the year will be extracted from the PDF file.
            
        operations : DataFrame, optional
            A DataFrame containing the operations. If not provided, the DataFrame will be initialized with an empty DataFrame.
            
        **kwargs : dict
            Additional keyword arguments to pass to the class.
            Examples: verbose (bool), do_log (bool, if True, logs will be saved to a file), formatting (bool, if True, logs will be formatted), rules_file (str, the path to the rules file), ask_rules (bool, if True, the user will be asked to provide rules for the categories), handle_errors (bool, if True, errors will be handled and logged)
            
        Returns:
        --------
        None
        """
        self.parse_args(kwargs)

        self.month = month
        self.year = year
        if operations is None:
            self.operations = pd.DataFrame(columns = ["Date", "Description", "Operation Date", "Debit (€)", "Credit (€)", "Category"])
        else:
            self.operations = operations
        self.pdf = pdf
        now = datetime.now()
        # format the date as YYYY-MM-DD_HH-MM-SS
        now = now.strftime("%Y-%m-%d_%H-%M-%S")
        logger_name = f"{LOG_FOLDER}/monthly_summary_{self.pdf.split('/')[-1].split('.')[0]}_{now}.log"
        self.logger = Logger(logger_name, do_log = self.do_log, verbose = self.verbose, formatting = self.formatting)
        self.logger.log(f"Monthly summary created for {month} {year} with {len(self.operations)} operations")
        self.logger.log(f"PDF file: {pdf}")
        
    def parse_args(self, kwargs):
        """Parse the keyword arguments and set the attributes of the class.

        Parameters:
        -----------
        kwargs : dict
            The keyword arguments to parse.

        Returns:
        --------
        None
        """

        for key, value in kwargs.items():
            setattr(self, key, value)

        if "verbose" not in kwargs and not hasattr(self, "verbose"):
            self.verbose = True
        if "do_log" not in kwargs and not hasattr(self, "do_log"):
            self.do_log = True
        if "formatting" not in kwargs and not hasattr(self, "formatting"):
            self.formatting = True
        if "rules_file" not in kwargs and not hasattr(self, "rules_file"):
            self.rules_file = None
        if "ask_rules" not in kwargs and not hasattr(self, "ask_rules"):
            self.ask_rules = False
        if "handle_errors" not in kwargs and not hasattr(self, "handle_errors"):
            self.handle_errors = True

        return None

    @timeit
    def add_operations(self, mode = "text", **kwargs):
        """Load the PDF file, parse the operations and add them to the DataFrame.

        Parameters:
        -----------
        mode : str, optional
            The mode to use to parse the PDF file. Can be "text" or "markdown". Default is "text".  
        **kwargs : dict
            Additional keyword arguments to pass to the class.
            Examples: verbose (bool), do_log (bool, if True, logs will be saved to a file), formatting (bool, if True, logs will be formatted), rules_file (str, the path to the rules file), ask_rules (bool, if True, the user will be asked to provide rules for the categories), handle_errors (bool, if True, errors will be handled and logged)

        Returns:
        --------
        operations : DataFrame
            The DataFrame containing the operations.
        """
        self.parse_args(kwargs)
        self.logger.verbose = self.verbose
        self.logger.do_log = self.do_log
        self.logger.formatting = self.formatting


        self.parser = Statement_Parser(self.pdf, mode=mode, logger = self.logger, **kwargs)
        self.parser.load_document()
        
        if self.parser.parsed_document is not None:
            self.operations = self.parser.parse_document()
            self.logger.log(f"Operations successfully added to the monthly summary for {self.month} {self.year}")


        try:
            self.operations = self.add_category(self.operations, ask_rules = self.ask_rules, rules_file = self.rules_file)
            self.logger.log(f"Categories added to the monthly summary for {self.month} {self.year}")
        except Exception as e:
            self.logger.error(f"Error adding categories to the monthly summary: {e}", title = "Category error")
            if not self.handle_errors:
                raise e
        
        # sort the operations by date
        self.operations = self.operations.sort_values(by="Date")

        self.month = self.operations["Date"].iloc[0].month
        self.year = self.operations["Date"].iloc[0].year
        self.start_date = self.parser.start_date
        self.end_date = self.parser.end_date

            
        if hasattr(self, "budget"):
            self.compute_remaining_budget()


        self.logger.log(f"Operations successfully processed. {len(self.operations)} operations found")
        return self.operations

    def modify_operation(self, index, column, value):
        """Modify an operation in the DataFrame.

        Parameters:
        -----------
        index : int
            The index of the operation to modify.

        column : str
            The column to modify.

        value : str or float
            The new value to assign to the column.

        Returns:
        --------
        None    
        """
        if self.operations is None:
            self.logger.warning("No operations to modify", title = "Modification warning")
            return None
        if index >= len(self.operations):
            self.logger.warning("Index out of range", title = "Modification warning")
            return None
        if column not in self.operations.columns:
            self.logger.warning("Column not found", title = "Modification warning")
            return None
        self.operations.at[index, column] = value
        self.logger.log(f"Operation {index} modified with value {value} in column {column}")
        return None

    def plot_categories(self, epargne = False):
        """Plot a pie chart of the expenses by categories.

        Parameters:
        -----------
        epargne : bool, optional
            If True, the "Epargne" category will be included in the pie chart. If False, it will be excluded.

        Returns:
        --------
        None
        """
        import matplotlib.pyplot as plt
        import numpy as np

        if len(self.operations) == 0:
            self.logger.warning("No operations to plot", title = "Plotting warning")
            return None
        
        # store the categories and their total amount
        categories = self.operations[["Category", "Debit (€)"]].groupby("Category").sum()
        if not epargne:
            categories = categories[categories.index != "Epargne"]
        categories = categories.sort_values(by="Debit (€)", ascending=False)
        categories = categories.reset_index()

        # plot the pie chart
        fig, ax = plt.subplots()
        ax.pie(categories["Debit (€)"], labels=categories["Category"], autopct='%1.1f%%')
        ax.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.

        plt.show()
        return None
    
    def get_stats(self, epargne = False, print_stats = True):
        """Get the statistics of the monthly summary.

        Parameters:
        -----------
        - epargne : bool, optional
            If True, the statistics will include the "Epargne" category. If False, it will be excluded.

        - print_stats : bool, optional
            If True, the statistics will be printed. If False, they will only be returned.

        Returns:
        --------
        - stats : dict
            A dictionary containing the statistics of the monthly summary: total debit, total credit, total balance, saving rate.
        """
        if len(self.operations) == 0:
            self.logger.warning("No operations to get statistics from", title = "Statistics warning")
            return None
        
        if epargne:
            total_debit = self.operations["Debit (€)"].sum()
        else:
            total_debit = self.operations[self.operations["Category"] != "Epargne"]["Debit (€)"].sum()

        total_credit = self.operations["Credit (€)"].sum()
        total_balance = total_credit - total_debit
        saving_rate = (self.operations[self.operations["Category"] == "Epargne"]["Debit (€)"].sum() +total_balance) / total_credit

        v = self.logger.verbose
        self.logger.verbose = 0
        self.logger.log(f"Total debit: {total_debit}")
        self.logger.log(f"Total credit: {total_credit}")
        self.logger.log(f"Total balance: {total_balance}")
        self.logger.log(f"Saving rate: {saving_rate}")
        self.logger.verbose = v

        if print_stats:
            print(f"Total debit: {total_debit}")
            print(f"Total credit: {total_credit}")
            print(f"Total balance: {total_balance}")
            print(f"Saving rate: {saving_rate}")

        return {"Total debit": total_debit, "Total credit": total_credit, "Total balance": total_balance, "Saving rate": saving_rate}
    
    def summary(self):
        """Print a summary of the monthly report."""
        print(f"Monthly report summary for {self.month} {self.year}")
        if self.operations is not None:
            print(f"Number of operations: {len(self.operations)}")
        print(f"Start date: {self.start_date} \t End date: {self.end_date}")
        if hasattr(self, "remaining_budget"):
            print(f"Budget: {self.budget:.2f} \t\t\tRemaining budget: {self.remaining_budget:.2f}")
        if len(self.operations) > 0:
            stats = self.get_stats(print_stats = False)
            print(f"Total debit: {stats['Total debit']:.2f} \t\tTotal credit: {stats['Total credit']:.2f}")
            print(f"Total balance: {stats['Total balance']:.2f} \t\tSaving rate: {stats['Saving rate']:.2f}")
        return None

    def add_monthly_budget(self, budget = None):
        """Add a budget to the monthly summary.

        Parameters:
        -----------
        budget : float, optional
            The budget to add to the monthly summary. If not provided, an automatic budget will be computed.

        Returns:
        --------
        None
        """
        if budget is None:
            self.logger.warning("No budget provided, using automatic budget", title = "Budget warning")
            # automatic budget is 150 + 5*number of days in the month
            if hasattr(self, "start_date") and hasattr(self, "end_date"):
                budget = 150 + 5*((self.end_date - self.start_date).days)
            else:
                self.logger.error("No start and end dates found, cannot compute automatic budget", title = "Budget error")
                return None

        self.budget = budget

        if self.operations is not None and len(self.operations) != 0:
            self.compute_remaining_budget()

        return None
    
    def compute_remaining_budget(self, print_budget = True):
        """Compute the remaining budget.
        """
        stats = self.get_stats(print_stats = False)
        self.remaining_budget = self.budget - stats["Total debit"]
        self.logger.log(f"Total budget for {months_mapping_inv[self.month]} {self.year}: {self.budget}")
        self.logger.log(f"Remaining budget: {self.remaining_budget}")
        if print_budget:
            print(f"Total budget for {months_mapping_inv[self.month]} {self.year}: {self.budget}. Remaining budget: {self.remaining_budget}")
        return self.remaining_budget

    def to_csv(self, file = None):
        """Save the monthly summary to a CSV file.

        Parameters:
        -----------
        file : str, optional
            The path to the CSV file. If not provided, the file will be named "ACCOUNT_ID_month_year.csv".

        Returns:
        --------
        None
        """
        if file is None:
            file = f"{ACCOUNT_ID}_{self.month}_{self.year}.csv"

        try:
            with open(file, "w") as f:
                f.write(f"Relevé de compte issu du document PDF {self.pdf}\n")
                f.write(f"Mois,Année\n")
                f.write(f"{self.month},{self.year}\n")
                f.write("\n")
                f.write("Date,Description,Date d'opération,Débit (€),Crédit (€),Catégorie\n")
                for i in range(len(self.operations)):
                    f.write(f"{self.operations.iloc[i]['Date']},{self.operations.iloc[i]['Description']},{self.operations.iloc[i]['Operation Date']},{self.operations.iloc[i]['Debit (€)']},{self.operations.iloc[i]['Credit (€)']},{self.operations.iloc[i]['Category']}\n")

            self.logger.log(f"CSV file {file} created")

        except Exception as e:
            self.logger.error(f"Error while writing the CSV file: {e}", title = "CSV writing error")
            if not self.handle_errors:
                raise e
            return None

        return None
    
    def to_excel(self, file = None):
        """Save the monthly summary to an Excel file, and format the data. Also add a pie chart of the categories.

        Parameters:
        -----------
        file : str, optional
            The path to the Excel file. If not provided, the file will be named "ACCOUNT_ID_month_year.xlsx".

        Returns:
        --------
        None
        """
        import openpyxl
        from openpyxl import load_workbook
        from openpyxl.chart import PieChart, Reference
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill
        from openpyxl.formatting.rule import CellIsRule

        if file is None:
            file = f"{ACCOUNT_ID}_{self.month}_{self.year}.xlsx"
        
        number_to_month = {1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril", 5: "Mai", 6: "Juin", 7: "Juillet", 8: "Août", 9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre"}

        sheet_name = f"{number_to_month[self.month]}_{self.year}"

        if not os.path.exists(file):
            # Create the Excel file with the first sheet
            try:
                with pd.ExcelWriter(file, engine='openpyxl') as writer:
                    writer.book.create_sheet(sheet_name, 0)
                    writer.sheets[sheet_name].title = "Sheet_0"
                    self.logger.log(f"Excel file {file} created")
            except Exception as e:
                self.logger.error(f"Error while creating the Excel file: {e}", title = "Excel creation error")
                if not self.handle_errors:
                    raise e
                return None

        if not hasattr(self, "budget"):
            self.logger.warning("No budget provided, computing automatic budget", title = "Budget warning")
            self.add_monthly_budget()

        try:
            with pd.ExcelWriter(file, engine='openpyxl', mode="a") as writer:
                # Write first an introductory phrase: "Comptes pour le mois de {month} {year}"
                # Skip a line, Then writes "Budget: {budget} €" and next cell "Remaining budget: {remaining_budget} €"
                # Skip a line, Then writes the operations DataFrame

                writer.book.create_sheet(sheet_name, 0)
                writer.sheets[sheet_name].title = sheet_name
                title_cell = writer.sheets[sheet_name].cell(row=1, column=1)
                title_cell.value = f"Comptes pour le mois de {number_to_month[self.month]} {self.year} (du {self.start_date.date()} au {self.end_date.date()})"
                title_cell.font = Font(bold=True, size=20)
                      
                # Adjust the height of the row
                writer.sheets[sheet_name].row_dimensions[1].height = 30

                budget_cell = writer.sheets[sheet_name].cell(row=2, column=1)
                budget_cell.value = "Budget:"
                budget_cell.font = Font(bold=True)
                writer.sheets[sheet_name].cell(row=2, column=2, value=self.budget)

                remaining_budget_cell = writer.sheets[sheet_name].cell(row=3, column=1)
                remaining_budget_cell.value = "Remaining budget:"
                remaining_budget_cell.font = Font(bold=True)
                remaining_budget_cell = writer.sheets[sheet_name].cell(row=3, column=2, value=self.remaining_budget)
                # set some conditional formatting to have red if remaining budget is negative and green if positive
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                writer.sheets[sheet_name].conditional_formatting.add(f"B3:B3", CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
                writer.sheets[sheet_name].conditional_formatting.add(f"B3:B3", CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))

                #self.operations.to_excel(writer, sheet_name=sheet_name, index=False)

                # Write the column names in bold
                for c_idx, value in enumerate(self.operations.columns, start=1):
                    writer.sheets[sheet_name].cell(row=5, column=c_idx, value=value)
                    writer.sheets[sheet_name].cell(row=5, column=c_idx).font = Font(bold=True, size=16, color="FFFFFF")
                    # Color the cell in black and the font in white
                    writer.sheets[sheet_name].cell(row=5, column=c_idx).fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

                # Set the right row height for the header row
                writer.sheets[sheet_name].row_dimensions[5].height = 18

                for r_idx, r in enumerate(dataframe_to_rows(self.operations, index=False, header=False), start=6):
                    for c_idx, value in enumerate(r, start=1):
                        # Format the date to dd/mm/yyyy
                        if c_idx in [1, 3]:
                            writer.sheets[sheet_name].cell(row=r_idx, column=c_idx, value=value).number_format = 'dd/mm/yyyy'
                        else:
                            writer.sheets[sheet_name].cell(row=r_idx, column=c_idx, value=value)

                # Add a "" | "Total" row at the end of the operations rows
                writer.sheets[sheet_name].cell(row=len(self.operations) + 6, column=2, value="Total")
                writer.sheets[sheet_name].cell(row=len(self.operations) + 6, column=2).font = Font(bold=True, color="FFFFFF")
                writer.sheets[sheet_name].cell(row=len(self.operations) + 6, column=4, value=f"=SUM(D6:D{len(self.operations) + 5})")
                writer.sheets[sheet_name].cell(row=len(self.operations) + 6, column=4).font = Font(bold=True, color="FFFFFF")
                writer.sheets[sheet_name].cell(row=len(self.operations) + 6, column=5, value=f"=SUM(E6:E{len(self.operations) + 5})")
                writer.sheets[sheet_name].cell(row=len(self.operations) + 6, column=5).font = Font(bold=True, color="FFFFFF")
                for c_idx in [0, 1, 2, 3, 4, 5]:
                    writer.sheets[sheet_name].cell(row=len(self.operations) + 6, column=c_idx + 1).fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")


                # Set the right column width
                for col, w in {0: 17, 1: 30, 2: 20, 3:15, 4:15, 5:15}.items():
                    writer.sheets[sheet_name].column_dimensions[chr(65 + col)].width = w

                # Add frame lines around the data
                for r in writer.sheets[sheet_name].iter_rows(min_row=5, max_row=len(self.operations) + 5, min_col=1, max_col=len(self.operations.columns)):
                    for cell in r:
                        cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))

                # Colour each row alternatively in very light blue and gray
                for r_idx in range(6, len(self.operations) + 6):
                    fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    if r_idx % 2 == 0:
                        fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    for c_idx in range(1, len(self.operations.columns) + 1):
                        writer.sheets[sheet_name].cell(row=r_idx, column=c_idx).fill = fill
            self.logger.log(f"Excel file {file} created")

        except Exception as e:
            self.logger.error(f"Error while writing the Excel file: {e}", title = "Excel writing error")
            if not self.handle_errors:
                raise e
            return None
        
        try:
            # Load the workbook and select the active worksheet
            wb = load_workbook(file)
            ws = wb[sheet_name]

            # Create a summary DataFrame for the pie chart
            categories = self.operations[["Category", "Debit (€)"]].groupby("Category").sum()
            categories = categories.sort_values(by="Debit (€)", ascending=False)
            categories = categories.reset_index()

            # Determine the starting column for the summary data
            start_col = len(self.operations.columns) + 2  # Two columns to the right of the initial table
            start_row = 1
            categories_list = CATEGORY_LIST

            # Add category labels and sum formulas for debits
            for idx, category in enumerate(categories_list):
                category_cell = ws.cell(row=start_row + idx + 1, column=start_col, value=f"Débit {category} (€)")
                formula_cell = ws.cell(row=start_row + idx + 1, column=start_col + 1)
                formula_cell.value = f"=SUMIF(F:F, \"{category}\", D:D)"  # Sum of debits for each category
                if category == "Epargne":
                    formula_cell.value = 0
                    ws.cell(row=start_row + idx + 1, column=start_col + 2, value = f"=SUMIF(F:F, \"{category}\", D:D)")

            # Add a frame around the summary data
            for r in ws.iter_rows(min_row=start_row+1, max_row=start_row + len(categories_list), min_col=start_col, max_col=start_col + 1):
                for cell in r:
                    cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))

            # Create a PieChart for debits
            pie_debit = PieChart()
            labels_debit = Reference(ws, min_col=start_col, min_row=start_row+1, max_row=start_row + len(categories_list))
            data_debit = Reference(ws, min_col=start_col + 1, min_row=start_row, max_row=start_row + len(categories_list))
            pie_debit.add_data(data_debit, titles_from_data=True)
            pie_debit.set_categories(labels_debit)
            pie_debit.title = "Debit (€) by Category"

            # Add the pie chart to the worksheet
            ws.add_chart(pie_debit, f'{chr(65 + start_col + 4)}2')  # Position the chart a few columns to the right of the summary data

            # Add category labels and sum formulas for credits 10 rows below the debit section
            credit_start_row = start_row + len(categories_list) + 3
            for idx, category in enumerate(categories_list):
                category_cell = ws.cell(row=credit_start_row + idx + 1, column=start_col, value=f"Crédit {category} (€)")
                formula_cell = ws.cell(row=credit_start_row + idx + 1, column=start_col + 1)
                formula_cell.value = f"=SUMIF(F:F, \"{category}\", E:E)"  # Sum of credits for each category

            # Add a frame around the summary data
            for r in ws.iter_rows(min_row=credit_start_row+1, max_row=credit_start_row + len(categories_list), min_col=start_col, max_col=start_col + 1):
                for cell in r:
                    cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))

            # Set the right column width for col H width 20
            ws.column_dimensions['H'].width = 20

            # Create a PieChart for credits
            pie_credit = PieChart()
            labels_credit = Reference(ws, min_col=start_col, min_row=credit_start_row+1, max_row=credit_start_row + len(categories_list))
            data_credit = Reference(ws, min_col=start_col+1, min_row=credit_start_row, max_row=credit_start_row + len(categories_list))
            pie_credit.add_data(data_credit, titles_from_data=True)
            pie_credit.set_categories(labels_credit)
            pie_credit.title = "Credit (€) by Category"

            # Add the pie chart to the worksheet
            ws.add_chart(pie_credit, f'{chr(65 + start_col + 4)}{credit_start_row + 4}')  # Position the chart a few columns to the right of the summary data

            # add the saving rate
            sav_rate_cell = ws.cell(row=credit_start_row + len(categories_list) + 3, column=start_col, value="Saving rate")
            sav_rate_formula = ws.cell(row=credit_start_row + len(categories_list) + 3, column=start_col + 1)
            sav_rate_formula.value = self.get_stats(print_stats = False)["Saving rate"]

            # Save the workbook
            wb.save(file)

            self.logger.log(f"Summary data and pie chart added to the Excel file")

        except Exception as e:
            self.logger.error(f"Error while adding the summary data and pie chart to the Excel file: {e}", title = "Excel chart error")
            if not self.handle_errors:
                raise e
            return None

        return None
    
    @classmethod
    def build_rules(self, rules_file = None):
        """Class method to build the rules from a file."""
        if rules_file is None:
            rules_file = f"{ACCOUNT_ID}_rules.txt"
            
        rules = []
        with open(rules_file, "r") as f:
            for line in f:
                # Add the {string: category} mapping to the rules list
                rule = line.strip().split(":")
                if rule[1] not in CATEGORY_LIST:
                    rule[1] = "Other"
                rules.append({rule[0]: rule[1]})

        return rules, rules_file

    def _parse_operations(self, doc, page):
        """Parse the operations from a parsed PDF document."""
        d = doc.text
        try:
            info = d.split("|---|---|---|---|---|")[-1]
            info = info.split("\n\n")[0]
            infos = info.split("\n")
        except Exception as e:
            self.logger.error(f"Error parsing the PDF {self.pdf} on page {page}: {e}", title = "PDF parsing error")
            if not self.handle_errors:
                raise e

        data = []
        keys = ["Date", "Description", "Operation Date", "Debit (€)", "Credit (€)"]
        for i in infos:
            d = i.split("|")[1:6]
            if len(d)==5 and d[0]!="":
                data.append(dict(zip(keys, d)))

        data = pd.DataFrame(data)
        return data

    def _add_category(self, label, debit, credit, rules, ask_rules, rules_file):
        """Add a category to a transaction based on the label and the amount."""
        for rule in rules:
            for key in rule.keys():
                if key in label:
                    return rule[key]

        if ask_rules:
            print(f"Label: {label}, Debit: {debit}, Credit: {credit}")
            print("Please enter the category for this transaction among the following:")
            print(CATEGORY_LIST)
            category = input()
            print("Please enter the label to recognize it:")
            label = input()

            if category not in CATEGORY_LIST:
                category = "Other"
            elif label != "":
                rules.append({label: category})
                with open(rules_file, "a") as f:
                    f.write(f"{label}:{category}\n")

            return category
            
        return "Other"

    def add_category(self, data, ask_rules = None, rules_file = None):
        """Add a category to the operations based on the rules file."""
        if rules_file is None:
            rules_file = self.rules_file
        else:
            self.rules_file = rules_file
        if ask_rules is None:
            ask_rules = self.ask_rules
        else:
            self.ask_rules = ask_rules
        rules, rules_file = Monthly_Summary.build_rules(rules_file=rules_file)
        data["Category"] = data.apply(lambda x: self._add_category(x["Description"], x["Debit (€)"], x["Credit (€)"], rules, ask_rules, rules_file), axis=1)
        return data

    def _load_pdf(self):
        """Load the PDF file and parse it to extract the operations."""
        from llama_parse import LlamaParse
        parser = LlamaParse( # can also be set in your env as LLAMA_CLOUD_API_KEY
            result_type="markdown",  # "markdown" and "text" are available
            num_workers=4,  # if multiple files passed, split in `num_workers` API calls
            verbose=True,
            language="en",  # Optionally you can define a language, default=en
        )

        # sync
        documents = parser.load_data(self.pdf)
        return documents

    def __str__(self):
        return f"{self.month} {self.year} ; {self.operations}"
    

############################################################################################

##################################### HELPER FUNCTIONS #####################################

############################################################################################

def file_to_excel(file, excel_file, **kwargs):
    """Process a PDF file and save the monthly summary to an Excel file.
    
    Parameters:
    -----------
    file : str
        The path to the PDF file to process.
        
    excel_file : str
        The path to the Excel file to save the monthly summary to.

    **kwargs : dict
        Additional keyword arguments to pass to the Monthly_Summary class.
        Examples: verbose (bool), do_log (bool, if True, logs will be saved to a file), formatting (bool, if True, logs will be formatted), rules_file (str, the path to the rules file), ask_rules (bool, if True, the user will be asked to provide rules for the categories), handle_errors (bool, if True, errors will be handled and logged)
    """
    ms = Monthly_Summary(file, **kwargs)
    ms.add_operations()
    ms.add_monthly_budget()
    ms.to_excel(excel_file)
    return None

def process_files(files, dest_file = None, **kwargs):
    """Process a list of PDF files and save the monthly summaries to an Excel file.

    Parameters:
    -----------
    files : list or str
        The list of PDF files to process, or a single file.

    dest_file : str, optional
        The path to the Excel file to save the monthly summaries to. If not provided, the file will be named "ACCOUNT_ID_month_year.xlsx".

    **kwargs : dict
        Additional keyword arguments to pass to the Monthly_Summary class.

    Returns:
    --------
    None
    """
    if isinstance(files, str):
        files = [files]

    for i, file in enumerate(files):
        print(f"Processing file {i+1}/{len(files)}")
        if dest_file is None:
            file_to_excel(file, **kwargs)
        else:
            file_to_excel(file, dest_file, **kwargs)

    print("All files processed")
    return None

def process_folder(folder = "Data", dest_file = None, **kwargs):
    """Process all the PDF files in a folder and save the monthly summaries to an Excel file.

    Parameters:
    -----------
    folder : str, optional
        The path to the folder containing the PDF files. Default is "Data".

    dest_file : str, optional
        The path to the Excel file to save the monthly summaries to. If not provided, the file will be named "ACCOUNT_ID_month_year.xlsx".

    **kwargs : dict
        Additional keyword arguments to pass to the Monthly_Summary class.
        Examples: verbose (bool), do_log (bool, if True, logs will be saved to a file), formatting (bool, if True, logs will be formatted), rules_file (str, the path to the rules file), ask_rules (bool, if True, the user will be asked to provide rules for the categories), handle_errors (bool, if True, errors will be handled and logged)

    Returns:
    --------
    None
    """
    if "/" in dest_file:
        os.makedirs("/".join(dest_file.split("/")[:-1]), exist_ok=True)
    files = [os.path.join(folder, f) for f in os.listdir(folder) if f.endswith(".pdf")]
    process_files(files, dest_file, **kwargs)
    return None



############################################################################################

##################################### MAIN FUNCTION ########################################

############################################################################################

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Process a PDF bank statement to extract the operations and categorize them")
    parser.add_argument("file", type=str, help="The PDF file to process")
    parser.add_argument("--excel", type=str, help="The Excel file to save the data to")
    parser.add_argument("--budget", type=float, help="The monthly budget to set")
    parser.add_argument("--rules", type=str, help="The file containing the rules to categorize the operations")
    parser.add_argument("--ask_rules", action="store_true", help="Ask for the category of each operation")
    parser.add_argument("--handle_errors", action="store_false", help="Handle errors in the code")
    parser.add_argument("--verbose", action="store_true", help="Print verbose messages")
    parser.add_argument("--do_log", action="store_true", help="Log the messages")
    parser.add_argument("--formatting", action="store_true", help="Format the log messages")

    args = parser.parse_args()

    ms = Monthly_Summary(args.file, verbose = args.verbose, do_log = args.do_log, formatting = args.formatting, rules_file = args.rules, ask_rules = args.ask_rules, handle_errors = args.handle_errors)
    ms.add_operations()
    if args.budget is not None:
        ms.add_monthly_budget(args.budget)
    if args.excel is not None:
        ms.to_excel(args.excel)
    return None

if __name__ == "__main__":
    main()

############################################################################################

##################################### END OF SCRIPT ########################################

############################################################################################